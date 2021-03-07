import openpyxl

path = 'Open-source-data.xlsx'
workbook = openpyxl.load_workbook(path)


def which_sheet(sheet=1, col=1):
    worksheet = workbook.active
    if sheet == 1:
        worksheet = workbook['Africa']
    if sheet == 2:
        worksheet = workbook['Australasia']
    if sheet == 3:
        worksheet = workbook['Europe']
    if sheet == 4:
        worksheet = workbook['Latin America']
    if sheet == 5:
        worksheet = workbook['Middle East and Arid Asia']
    if sheet == 6:
        worksheet = workbook['North America']
    if sheet == 7:
        worksheet = workbook['Small Island States']
    if sheet == 8:
        worksheet = workbook['Temperate Asia']
    if sheet == 9:
        worksheet = workbook['Tropical Asia']

    i = 2
    k = col
    elements = []
    line = True
    while line:
        element = worksheet.cell(row=i, column=k)
        if element.value is None:
            line = False
        else:
            elements.append(element.value)
            i = i + 1
    return elements


def index_num(sheet=1, col=1):
    worksheet = workbook.active
    if sheet == 1:
        worksheet = workbook['Africa']
    if sheet == 2:
        worksheet = workbook['Australasia']
    if sheet == 3:
        worksheet = workbook['Europe']
    if sheet == 4:
        worksheet = workbook['Latin America']
    if sheet == 5:
        worksheet = workbook['Middle East and Arid Asia']
    if sheet == 6:
        worksheet = workbook['North America']
    if sheet == 7:
        worksheet = workbook['Small Island States']
    if sheet == 8:
        worksheet = workbook['Temperate Asia']
    if sheet == 9:
        worksheet = workbook['Tropical Asia']

    i = 2
    k = col
    row_length = 0
    line = True
    while line:
        element = worksheet.cell(row=i, column=k)
        if element.value is None:
            line = False
        else:
            row_length = row_length + 1
            i = i + 1
    return row_length


def index_array(x, y):
    indexes = []
    index = 0
    for i in range(index_num(x, y)):
        indexes.append(index)
        index = index + 1
    return indexes
